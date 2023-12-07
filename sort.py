import openpyxl
import re

def sanitize_filename(filename):
    return re.sub(r'[<>:"/\\|?*]', '_', filename)

file_path = "schools.xlsx" # file path

workbook = openpyxl.load_workbook(file_path)

sheet = workbook.active
header = sheet[1] # if header in first row
header_values = [cell.value for cell in header]

# create a dictionary mapping schools to their students
students_by_school = {}

# populates dictionary with schools and corresponding students
for row in sheet.iter_rows(min_row=2, values_only=True):
    student_school = row[1]  # Assuming row 1 contains school id/name
    
    # If school not encounterd yet, add new
    if student_school not in students_by_school:
        students_by_school[student_school] = []
    
    # Add student to their school
    students_by_school[student_school].append(row)

# Creates new workbooks for each school with students as entries
for school, students in students_by_school.items():
    new_workbook = openpyxl.Workbook()
    new_sheet = new_workbook.active

    new_sheet.append(header_values)

    for student_data in students:
        new_sheet.append(student_data)

    # name the workbook as the school name
    sanitized_school_name = sanitize_filename(school)
    new_file_name = f"{sanitized_school_name}_students.xlsx"

    # save the new workbook
    new_workbook.save(new_file_name)
    new_workbook.close()

workbook.close()