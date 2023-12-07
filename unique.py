import openpyxl
import re

def sanitize_filename(filename):
    return re.sub(r'[<>:"/\\|?*]', '_', filename)

file_path = "schools.xlsx"
new_file_name = "unique_schools.xlsx"

workbook = openpyxl.load_workbook(file_path)
sheet = workbook.active
header = sheet[1]
header_values = [cell.value for cell in header]

students_by_school = {}

# Using enumerate to keep track of the row number
for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
    student_school = row[5]  # Assuming school information is in column 5
    
    # Check if the school is not already in the dictionary
    if student_school not in students_by_school:
        students_by_school[student_school] = row  # Store the row data

new_workbook = openpyxl.Workbook()
new_sheet = new_workbook.active

new_sheet.append(header_values)  # Add the header to the new sheet

# Adding unique entries to the new sheet
for school, student_data in students_by_school.items():
    new_sheet.append(student_data)

new_workbook.save(new_file_name)
new_workbook.close()
workbook.close()
